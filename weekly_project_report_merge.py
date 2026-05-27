from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

INPUT_DIR = Path("input")
OUTPUT_FILE = Path("output/weekly_project_summary.xlsx")

OUTPUT_HEADERS = [
    "来源文件",
    "来源工作表",
    "序号",
    "项目名称",
    "项目概况",
    "执行人员",
    "项目进展情况",
    "本周工作进展",
    "下一步工作",
    "下一步工作时间节点",
]

FIELD_ALIASES: Dict[str, List[str]] = {
    "序号": ["序号", "编号"],
    "项目名称": ["项目名称", "项目名"],
    "项目概况": ["项目概况", "项目简介", "项目说明"],
    "执行人员": ["执行人员", "执行人", "负责人"],
    "项目进展情况": ["项目进展情况", "项目进展", "进展情况"],
    "本周工作进展": ["本周工作进展", "本周进展", "本周工作"],
    "下一步工作": ["下一步工作", "后续工作", "下周工作"],
    "下一步工作时间节点": ["下一步工作时间节点", "时间节点", "下一步时间"],
}


@dataclass
class Stats:
    excel_files: int = 0
    worksheets: int = 0
    records: int = 0
    skipped_blank_rows: int = 0


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_for_match(value: object) -> str:
    text = normalize_text(value)
    return "".join(text.split())


def fill_merged_cells(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)


def match_field(cell_value: object) -> Optional[str]:
    normalized = normalize_for_match(cell_value)
    if not normalized:
        return None

    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = "".join(alias.split())
            if key and (key in normalized or normalized in key):
                return field_name
    return None


def detect_header(ws, scan_limit: int = 50) -> tuple[Optional[int], Dict[str, int]]:
    best_row = None
    best_map: Dict[str, int] = {}

    max_row = min(ws.max_row, scan_limit)
    for row_idx in range(1, max_row + 1):
        current_map: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            field = match_field(ws.cell(row=row_idx, column=col_idx).value)
            if field and field not in current_map:
                current_map[field] = col_idx

        score = len(current_map)
        if score > len(best_map):
            best_map = current_map
            best_row = row_idx

    required = {"项目名称", "本周工作进展", "下一步工作"}
    if best_row is None or not required.issubset(best_map.keys()):
        return None, {}
    return best_row, best_map


def is_blank_record(record: Dict[str, str]) -> bool:
    return all(not normalize_text(v) for v in record.values())


def parse_sheet(ws, source_file: str, stats: Stats) -> List[Dict[str, str]]:
    fill_merged_cells(ws)
    header_row, col_map = detect_header(ws)
    if header_row is None:
        return []

    records: List[Dict[str, str]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {field: "" for field in FIELD_ALIASES.keys()}
        for field, col_idx in col_map.items():
            row_data[field] = normalize_text(ws.cell(row=row_idx, column=col_idx).value)

        if is_blank_record(row_data):
            stats.skipped_blank_rows += 1
            continue

        summary_row = {
            "来源文件": source_file,
            "来源工作表": ws.title,
            "序号": row_data.get("序号", ""),
            "项目名称": row_data.get("项目名称", ""),
            "项目概况": row_data.get("项目概况", ""),
            "执行人员": row_data.get("执行人员", ""),
            "项目进展情况": row_data.get("项目进展情况", ""),
            "本周工作进展": row_data.get("本周工作进展", ""),
            "下一步工作": row_data.get("下一步工作", ""),
            "下一步工作时间节点": row_data.get("下一步工作时间节点", ""),
        }
        records.append(summary_row)

    return records


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.split("\n"):
                max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def write_output(records: List[Dict[str, str]]) -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "周报汇总"

    ws.append(OUTPUT_HEADERS)
    for row in records:
        ws.append([row.get(header, "") for header in OUTPUT_HEADERS])

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    wb.save(OUTPUT_FILE)


def main() -> None:
    stats = Stats()
    all_records: List[Dict[str, str]] = []

    files = sorted(
        p for p in INPUT_DIR.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")
    )
    stats.excel_files = len(files)

    for file_path in files:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            stats.worksheets += 1
            rows = parse_sheet(ws, file_path.name, stats)
            all_records.extend(rows)
        wb.close()

    stats.records = len(all_records)
    write_output(all_records)

    print(f"读取 Excel 文件数: {stats.excel_files}")
    print(f"读取工作表数: {stats.worksheets}")
    print(f"成功汇总项目记录数: {stats.records}")
    print(f"跳过空白行数: {stats.skipped_blank_rows}")
    print(f"输出文件: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
