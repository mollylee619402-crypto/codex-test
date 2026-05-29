from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "project_file_inventory.xlsx"

OUTPUT_HEADERS = [
    "序号",
    "文件名",
    "扩展名",
    "文件类型",
    "所在文件夹",
    "完整路径",
    "文件大小KB",
    "修改时间",
    "备注",
]

FILE_TYPE_MAP = {
    ".doc": "Word 文档",
    ".docx": "Word 文档",
    ".xls": "Excel 表格",
    ".xlsx": "Excel 表格",
    ".pdf": "PDF 文档",
    ".dwg": "CAD 图纸",
    ".dxf": "DXF 图纸",
    ".csv": "CSV 表格",
    ".txt": "文本文件",
    ".jpg": "图片",
    ".jpeg": "图片",
    ".png": "图片",
    ".zip": "压缩包",
    ".rar": "压缩包",
    ".7z": "压缩包",
}

COLUMN_WIDTHS = {
    "A": 8,
    "B": 28,
    "C": 12,
    "D": 16,
    "E": 36,
    "F": 56,
    "G": 14,
    "H": 22,
    "I": 24,
}


@dataclass
class FileInventoryItem:
    index: int
    name: str
    extension: str
    file_type: str
    folder: str
    full_path: str
    size_kb: float
    modified_time: str
    note: str = ""

    def to_row(self) -> List[object]:
        return [
            self.index,
            self.name,
            self.extension,
            self.file_type,
            self.folder,
            self.full_path,
            self.size_kb,
            self.modified_time,
            self.note,
        ]


def ensure_directories() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def get_file_type(extension: str) -> str:
    return FILE_TYPE_MAP.get(extension.lower(), "其他文件")


def format_modified_time(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def scan_input_files() -> List[FileInventoryItem]:
    files = sorted(path for path in INPUT_DIR.rglob("*") if path.is_file())
    inventory_items: List[FileInventoryItem] = []

    for index, file_path in enumerate(files, start=1):
        stat = file_path.stat()
        extension = file_path.suffix.lower()
        inventory_items.append(
            FileInventoryItem(
                index=index,
                name=file_path.name,
                extension=extension,
                file_type=get_file_type(extension),
                folder=str(file_path.parent.resolve()),
                full_path=str(file_path.resolve()),
                size_kb=round(stat.st_size / 1024, 2),
                modified_time=format_modified_time(stat.st_mtime),
            )
        )

    return inventory_items


def apply_basic_style(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = body_alignment

    for column_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_inventory_excel(inventory_items: List[FileInventoryItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "项目资料清单"

    ws.append(OUTPUT_HEADERS)
    for item in inventory_items:
        ws.append(item.to_row())

    apply_basic_style(ws)
    wb.save(OUTPUT_FILE)


def main() -> None:
    ensure_directories()
    inventory_items = scan_input_files()
    write_inventory_excel(inventory_items)

    print(f"扫描到 {len(inventory_items)} 个文件")
    print(f"输出文件路径: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
